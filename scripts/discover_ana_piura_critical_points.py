#!/usr/bin/env python3
"""Descubre fichas ANA 2026 relevantes para Catacaos/Bajo Piura.

Procesa temporalmente ZIP oficiales SIGRID. Usa la capa de texto de PDFs para
clasificar fichas y, de preferencia, los anexos Excel para recuperar metadatos
estructurados. No usa OCR, no republica documentos y no convierte referencias
en alertas, umbrales ni coordenadas operativas sin validación explícita.
"""
from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
import json
import re
import tempfile
import zipfile

import requests
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from pypdf import PdfReader

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "site/data/hydrology/ana_piura_critical_points_2026.json"
MAX_DOWNLOAD = 120 * 1024 * 1024

DOCUMENTS = [
    {"id": 21411, "url": "https://sigrid.cenepred.gob.pe/sigridv3/documento/21411/descargar?boletin=578", "office": "0127-2026-ANA-J"},
    {"id": 22176, "url": "https://sigrid.cenepred.gob.pe/sigridv3/documento/22176/descargar", "office": "0774-2026-ANA-J"},
    {"id": 22191, "url": "https://sigrid.cenepred.gob.pe/sigridv3/documento/22191/descargar", "office": "0788-2026-ANA-J"},
]

TERRITORY = {
    "catacaos": ["catacaos"],
    "la_legua": ["la legua"],
    "simbila": ["simbilá", "simbila"],
    "pedregal_grande": ["pedregal grande"],
    "cura_mori": ["cura mori"],
    "bajo_piura": ["bajo piura"],
    "rio_piura": ["río piura", "rio piura"],
}
TARGET_ROW_TERMS = {k: v for k, v in TERRITORY.items() if k not in {"bajo_piura", "rio_piura"}}

INTERVENTIONS = {
    "descolmatacion": ["descolmatación", "descolmatacion"],
    "enrocado": ["enrocado"],
    "dique": ["dique"],
    "defensa_riberena": ["defensa ribereña", "defensa riberena"],
    "limpieza_cauce": ["limpieza de cauce", "limpieza del cauce"],
    "encauzamiento": ["encauzamiento"],
    "reforestacion": ["reforestación", "reforestacion"],
}
HAZARDS = {
    "inundacion": ["inundación", "inundacion"],
    "erosion_fluvial": ["erosión fluvial", "erosion fluvial"],
    "desborde": ["desborde"],
    "socavacion": ["socavación", "socavacion"],
}
HEADER_WORDS = (
    "ficha", "departamento", "provincia", "distrito", "sector", "localidad",
    "centro poblado", "cauce", "río", "rio", "quebrada", "coordenada",
    "este", "norte", "easting", "northing", "latitud", "longitud", "tramo",
)


def clean_name(name):
    return name.replace("\\", "/").split("/")[-1]


def normalize(value):
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def download(url, headers):
    r = requests.get(url, timeout=(20, 150), headers=headers)
    r.raise_for_status()
    data = r.content
    if len(data) > MAX_DOWNLOAD:
        raise RuntimeError(f"Descarga excede límite de {MAX_DOWNLOAD} bytes")
    return data, r.headers.get("content-type", "")


def tags(text, groups):
    low = text.lower()
    return sorted(k for k, needles in groups.items() if any(n in low for n in needles))


def territory_hits(text, groups=TERRITORY):
    low = text.lower()
    return sorted(k for k, needles in groups.items() if any(n in low for n in needles))


def crs_tokens(text):
    values = []
    checks = [
        (r"(?i)WGS\s*[- ]?84", "WGS84"),
        (r"(?i)UTM", "UTM"),
        (r"(?i)(?:zona|zone)\s*17\s*S", "UTM zone 17S"),
        (r"(?i)17\s*S", "17S"),
    ]
    for pattern, label in checks:
        if re.search(pattern, text) and label not in values:
            values.append(label)
    return values


def labeled_values(raw, label, maxlen=120):
    values = []
    for pat in (
        rf"(?im)^\s*{label}\s*[:\-]?\s*([^\n]{{2,{maxlen}}})",
        rf"(?i){label}\s*[:\-]\s*([^;,.]{{2,{maxlen}}})",
    ):
        for m in re.finditer(pat, raw):
            value = normalize(m.group(1)).strip(" :-–—,.;")
            if value and value.lower() not in {x.lower() for x in values}:
                values.append(value)
    return values[:8]


def raw_flow_tokens(text):
    # Se conserva el token original para no confundir 3.468 con 3.468 o 3468.
    return list(dict.fromkeys(
        m.group(1) for m in re.finditer(r"(\d{1,5}(?:[.,]\d{1,3})?)\s*(?:m3/s|m³/s)", text, re.I)
    ))[:20]


def extract_pdf(path):
    try:
        reader = PdfReader(str(path))
    except Exception:
        return {"page_count": 0, "text_layer_available": False, "territory_page_hits": {}, "structured": {}}

    page_hits = {k: [] for k in TERRITORY}
    raw_pages = []
    for pageno, page in enumerate(reader.pages, start=1):
        try:
            raw = page.extract_text() or ""
        except Exception:
            raw = ""
        if not raw.strip():
            continue
        raw_pages.append(raw)
        low = raw.lower()
        for key, needles in TERRITORY.items():
            if any(n in low for n in needles):
                page_hits[key].append(pageno)

    raw = "\n".join(raw_pages)
    text = normalize(raw)
    structured = {
        "district_candidates": labeled_values(raw, r"distrito"),
        "province_candidates": labeled_values(raw, r"provincia"),
        "sector_candidates": (
            labeled_values(raw, r"sector") +
            labeled_values(raw, r"centro\s+poblado") +
            labeled_values(raw, r"localidad")
        )[:12],
        "hazard_tags": tags(text, HAZARDS),
        "intervention_tags": tags(text, INTERVENTIONS),
        "crs_tokens": crs_tokens(text),
        "flow_value_tokens_unvalidated": raw_flow_tokens(text),
    }
    structured = {k: v for k, v in structured.items() if v}
    return {
        "page_count": len(reader.pages),
        "text_layer_available": bool(raw_pages),
        "territory_page_hits": {k: v for k, v in page_hits.items() if v},
        "structured": structured,
    }


def unique_headers(values):
    result = []
    counts = {}
    for idx, value in enumerate(values, start=1):
        base = normalize(value) or get_column_letter(idx)
        counts[base] = counts.get(base, 0) + 1
        result.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return result


def header_score(values):
    text = " | ".join(normalize(v).lower() for v in values if v is not None)
    return sum(1 for word in HEADER_WORDS if word in text)


def numeric(value):
    if isinstance(value, (int, float)):
        return float(value)
    text = normalize(value).replace(" ", "")
    if re.fullmatch(r"-?\d+(?:[.,]\d+)?", text):
        try:
            return float(text.replace(",", "."))
        except Exception:
            pass
    return None


def coordinate_fields(fields):
    out = {"crs_tokens": []}
    for key, value in fields.items():
        low = key.lower()
        text = normalize(value)
        for token in crs_tokens(f"{key} {text}"):
            if token not in out["crs_tokens"]:
                out["crs_tokens"].append(token)
        n = numeric(value)
        if n is None:
            continue
        if any(x in low for x in ("este", "easting", "utm e")) and 100000 <= n <= 900000:
            out["easting"] = n
        if any(x in low for x in ("norte", "northing", "utm n")) and 8000000 <= n <= 10000000:
            out["northing"] = n
        if "latitud" in low and -90 <= n <= 90:
            out["latitude"] = n
        if "longitud" in low and -180 <= n <= 180:
            out["longitude"] = n
    if not out["crs_tokens"]:
        out.pop("crs_tokens")
    return out


def parse_xlsx(raw, doc, member_name):
    rows_out = []
    try:
        wb = load_workbook(BytesIO(raw), read_only=True, data_only=True)
    except Exception as exc:
        return {"file_name": clean_name(member_name), "status": "error", "error_type": type(exc).__name__, "error": str(exc), "relevant_rows": []}

    for ws in wb.worksheets:
        rows = []
        for r_idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            values = list(row[:50])
            if any(v not in (None, "") for v in values):
                rows.append((r_idx, values))
            if r_idx >= 400:
                break
        if not rows:
            continue

        header_candidates = rows[:40]
        header_row_num, header_values = max(header_candidates, key=lambda item: header_score(item[1]))
        score = header_score(header_values)
        headers = unique_headers(header_values) if score >= 2 else None

        for row_num, values in rows:
            if row_num <= header_row_num:
                continue
            row_text = " | ".join(normalize(v) for v in values if v not in (None, ""))
            hits = territory_hits(row_text, TARGET_ROW_TERMS)
            if not hits:
                continue
            if headers:
                fields = {
                    headers[i]: normalize(values[i])[:180]
                    for i in range(min(len(headers), len(values)))
                    if values[i] not in (None, "")
                }
            else:
                fields = {
                    get_column_letter(i + 1): normalize(v)[:180]
                    for i, v in enumerate(values)
                    if v not in (None, "")
                }
            rows_out.append({
                "source_document_id": doc["id"],
                "office": doc["office"],
                "source_page": f"https://sigrid.cenepred.gob.pe/sigridv3/documento/{doc['id']}",
                "file_name": clean_name(member_name),
                "sheet": ws.title,
                "row_number": row_num,
                "target_hits": hits,
                "header_row_number": header_row_num if headers else None,
                "fields": fields,
                "coordinate_candidates": coordinate_fields(fields),
                "coordinates_validated": False,
                "production_use": False,
            })
    return {"file_name": clean_name(member_name), "status": "processed", "relevant_rows": rows_out}


def process_pdf(path, doc, member_name=None):
    parsed = extract_pdf(path)
    name = member_name or path.name
    name_hits = territory_hits(name)
    hits = parsed["territory_page_hits"]
    return {
        "source_document_id": doc["id"],
        "office": doc["office"],
        "source_page": f"https://sigrid.cenepred.gob.pe/sigridv3/documento/{doc['id']}",
        "file_name": clean_name(name),
        "page_count": parsed["page_count"],
        "text_layer_available": parsed["text_layer_available"],
        "territory_hits_from_filename": name_hits,
        "territory_page_hits": hits,
        "structured": parsed["structured"],
        "relevant_to_catacaos_model": bool(hits or name_hits),
        "production_use": False,
    }


def main():
    headers = {"User-Agent": "Mozilla/5.0 IRFEN-research/0.8"}
    report = {
        "version": "0.8-experimental",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "production_use": False,
        "status": "starting",
        "purpose": "Descubrir fichas ANA 2026 relacionadas con Catacaos/Bajo Piura y priorizar anexos estructurados para sector/coordenadas.",
        "documents": [],
        "relevant_files": [],
        "spreadsheet_relevant_rows": [],
        "warning": "Coordenadas y números extraídos son candidatos hasta validar CRS, campo y significado. No se usan en producción.",
    }

    with tempfile.TemporaryDirectory(prefix="irfen_ana_piura_") as td:
        work = Path(td)
        for doc in DOCUMENTS:
            item = {"document_id": doc["id"], "office": doc["office"], "download_url": doc["url"]}
            try:
                data, content_type = download(doc["url"], headers)
                item.update({"download_bytes": len(data), "content_type": content_type, "files": []})

                if zipfile.is_zipfile(BytesIO(data)):
                    item["container"] = "zip"
                    with zipfile.ZipFile(BytesIO(data)) as z:
                        members = [m for m in z.namelist() if not m.endswith("/")]
                        item["member_count"] = len(members)
                        for idx, member in enumerate(members):
                            name = clean_name(member)
                            ext = Path(name).suffix.lower()
                            item["files"].append({"file_name": name, "extension": ext, "territory_hits_from_filename": territory_hits(name)})
                            raw = z.read(member)
                            if ext == ".pdf":
                                p = work / f"{doc['id']}_{idx}.pdf"
                                p.write_bytes(raw)
                                result = process_pdf(p, doc, member)
                                if result["relevant_to_catacaos_model"]:
                                    report["relevant_files"].append(result)
                            elif ext in {".xlsx", ".xlsm"}:
                                sheet_result = parse_xlsx(raw, doc, member)
                                report["spreadsheet_relevant_rows"].extend(sheet_result.get("relevant_rows", []))
                elif data[:4] == b"%PDF":
                    item["container"] = "pdf"
                    p = work / f"{doc['id']}.pdf"
                    p.write_bytes(data)
                    result = process_pdf(p, doc)
                    if result["relevant_to_catacaos_model"]:
                        report["relevant_files"].append(result)
                else:
                    item.update({"container": "unknown", "error": "Formato descargado no reconocido como ZIP/PDF"})
            except Exception as exc:
                item.update({"status": "error", "error_type": type(exc).__name__, "error": str(exc)})
            else:
                item["status"] = "processed"
            report["documents"].append(item)

    report["relevant_file_count"] = len(report["relevant_files"])
    report["spreadsheet_relevant_row_count"] = len(report["spreadsheet_relevant_rows"])
    report["territory_summary"] = {
        key: sum(
            1 for f in report["relevant_files"]
            if key in f.get("territory_page_hits", {}) or key in f.get("territory_hits_from_filename", [])
        ) + sum(1 for r in report["spreadsheet_relevant_rows"] if key in r.get("target_hits", []))
        for key in TERRITORY
    }
    report["status"] = "discovered_with_structured_index" if report["spreadsheet_relevant_rows"] else ("discovered" if report["relevant_files"] else "processed_no_catacaos_match_yet")

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    print(json.dumps({
        "status": report["status"],
        "relevant_file_count": report["relevant_file_count"],
        "spreadsheet_relevant_row_count": report["spreadsheet_relevant_row_count"],
        "territory_summary": report["territory_summary"],
        "spreadsheet_rows": report["spreadsheet_relevant_rows"],
    }, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
